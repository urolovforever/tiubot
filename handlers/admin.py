from aiogram import types, Dispatcher
from aiogram.dispatcher import FSMContext
from keyboards.reply import (get_admin_keyboard, get_cancel_keyboard, get_events_keyboard,
                             get_main_keyboard, get_statistics_keyboard,
                             get_skip_keyboard, get_faculty_keyboard, get_course_keyboard,
                             get_direction_keyboard, get_group_keyboard)
# get_broadcast_confirm_keyboard endi ishlatilmaydi
from database.db import Database, get_tashkent_now
from states.forms import AdminReplyState, EventCreateState, EventDeleteState, ScheduleUploadState
# BroadcastState endi universal_broadcast.py da ishlatiladi
from utils.helpers import t, is_admin
from datetime import datetime
from config import ADMIN_IDS
import logging
import asyncio

db = Database()
logger = logging.getLogger(__name__)


# ==================== ADMIN PANEL ASOSIY ====================

async def admin_panel_handler(message: types.Message, state: FSMContext):
    user_id = message.from_user.id

    if not is_admin(user_id):
        return

    await state.finish()

    lang = db.get_user_language(user_id)
    texts = {
        'uz': '👨‍💼 Admin panel\n\nKerakli bo\'limni tanlang:',
        'ru': '👨‍💼 Админ панель\n\nВыберите нужный раздел:',
        'en': '👨‍💼 Admin Panel\n\nChoose the section:'
    }

    await message.answer(
        texts.get(lang, texts['uz']),
        reply_markup=get_admin_keyboard(user_id)
    )


# ==================== YANGI MUROJAATLAR ====================

async def view_new_applications_handler(message: types.Message):
    user_id = message.from_user.id

    if not is_admin(user_id):
        return

    applications = db.get_new_applications()

    if not applications:
        lang = db.get_user_language(user_id)
        texts = {
            'uz': '📭 Yangi murojaatlar yo\'q',
            'ru': '📭 Новых обращений нет',
            'en': '📭 No new applications'
        }
        await message.answer(
            texts.get(lang, texts['uz']),
            reply_markup=get_admin_keyboard(user_id)
        )
        return

    await message.answer(f'📬 Yangi murojaatlar: {len(applications)} ta\n\n')

    for app in applications:
        text = f'''📬 Murojaat #{app[0]}

👤 Foydalanuvchi:
  • Ism: {app[3]}
  • Username: @{app[2] if app[2] else "yo'q"}
  • Telefon: {app[4]}
  • Link: tg://user?id={app[1]}

💬 Murojaat:
{app[5]}

📅 Sana: {app[8]}

🆕 Status: Yangi

📌 Javob berish: /reply_{app[0]}'''

        if app[6]:  # file_id
            try:
                await message.answer_photo(app[6], caption=text)
            except:
                try:
                    await message.answer_document(app[6], caption=text)
                except:
                    await message.answer(text)
        else:
            await message.answer(text)

        await asyncio.sleep(0.3)  # Spam oldini olish


# ==================== JAVOB BERILGAN MUROJAATLAR ====================

async def view_answered_applications_handler(message: types.Message):
    user_id = message.from_user.id

    if not is_admin(user_id):
        return

    # Oxirgi 7 kunlik murojaatlarni olish
    applications = db.get_answered_applications(days=7)

    if not applications:
        lang = db.get_user_language(user_id)
        texts = {
            'uz': '📭 Oxirgi 7 kunda javob berilgan murojaatlar yo\'q',
            'ru': '📭 Нет отвеченных обращений за последние 7 дней',
            'en': '📭 No answered applications in the last 7 days'
        }
        await message.answer(
            texts.get(lang, texts['uz']),
            reply_markup=get_admin_keyboard(user_id)
        )
        return

    await message.answer(f'✅ Javob berilgan murojaatlar (oxirgi 7 kun): {len(applications)} ta\n\n')

    for app in applications[:20]:  # Oxirgi 20 ta
        text = f'''📬 Murojaat #{app[0]}
✅ Status: Javob berilgan

👤 Foydalanuvchi:
  • Ism: {app[3]}
  • Telefon: {app[4]}
  • ID: {app[1]}

💬 Murojaat:
{app[5]}

💬 Javob:
{app[9] if app[9] else "Yo'q"}

📅 Sana: {app[8]}'''

        await message.answer(text)
        await asyncio.sleep(0.3)


# ==================== MUROJAATGA JAVOB BERISH ====================

async def reply_command_handler(message: types.Message, state: FSMContext):
    user_id = message.from_user.id

    if not is_admin(user_id):
        return

    try:
        app_id = int(message.text.split('_')[1])

        # Murojaat mavjudligini tekshirish
        app = db.get_application(app_id)
        if not app:
            await message.answer('❌ Murojaat topilmadi')
            return

        await state.update_data(app_id=app_id)

        lang = db.get_user_language(user_id)
        texts = {
            'uz': f'💬 Murojaat #{app_id} uchun javob yozing:',
            'ru': f'💬 Напишите ответ для обращения #{app_id}:',
            'en': f'💬 Write reply for application #{app_id}:'
        }

        await message.answer(
            texts.get(lang, texts['uz']),
            reply_markup=get_cancel_keyboard(user_id)
        )
        await AdminReplyState.waiting_for_reply.set()
    except Exception as e:
        logger.error(f'Error in reply command: {e}')
        await message.answer('❌ Xatolik yuz berdi')


async def process_admin_reply(message: types.Message, state: FSMContext):
    user_id = message.from_user.id

    if message.text in ['❌ Bekor qilish', '❌ Отмена', '❌ Cancel']:
        await state.finish()
        await message.answer(
            '❌ Bekor qilindi',
            reply_markup=get_admin_keyboard(user_id)
        )
        return

    data = await state.get_data()
    app_id = data.get('app_id')
    response = message.text

    app = db.get_application(app_id)

    if app:
        # Update database
        db.update_application_response(app_id, response)

        # Foydalanuvchiga xabar yuborish
        user_lang = db.get_user_language(app[1])

        response_texts = {
            'uz': f'''✅ Murojaatingizga javob keldi!

📬 Sizning murojaat #{app_id}:
{app[5]}

💬 Javob:
{response}

📅 {get_tashkent_now().strftime("%Y-%m-%d %H:%M")}''',

            'ru': f'''✅ Получен ответ на ваше обращение!

📬 Ваше обращение #{app_id}:
{app[5]}

💬 Ответ:
{response}

📅 {get_tashkent_now().strftime("%Y-%m-%d %H:%M")}''',

            'en': f'''✅ Response received for your application!

📬 Your application #{app_id}:
{app[5]}

💬 Response:
{response}

📅 {get_tashkent_now().strftime("%Y-%m-%d %H:%M")}'''
        }

        # Foydalanuvchiga va adminga xabarlarni parallel yuborish
        try:
            results = await asyncio.gather(
                message.bot.send_message(
                    app[1],
                    response_texts.get(user_lang, response_texts['uz'])
                ),
                message.answer(
                    f'✅ Javob yuborildi!\n\nMurojaat #{app_id} holati "Javob berilgan" ga o\'zgartirildi',
                    reply_markup=get_admin_keyboard(user_id)
                ),
                return_exceptions=True
            )

            # Agar foydalanuvchiga yuborishda xatolik bo'lsa
            if isinstance(results[0], Exception):
                logger.error(f'Error sending response to user: {results[0]}')
                await message.answer(
                    f'⚠️ Ogohlantirish: Foydalanuvchiga xabar yuborilmadi ({results[0]})\n\nLekin javob database\'ga saqlandi',
                    reply_markup=get_admin_keyboard(user_id)
                )
        except Exception as e:
            logger.error(f'Error in parallel operations: {e}')
            await message.answer(
                f'❌ Xatolik: {e}',
                reply_markup=get_admin_keyboard(user_id)
            )
    else:
        await message.answer(
            '❌ Murojaat topilmadi',
            reply_markup=get_admin_keyboard(user_id)
        )

    await state.finish()


# ==================== STATISTIKA ====================

async def statistics_handler(message: types.Message):
    user_id = message.from_user.id

    if not is_admin(user_id):
        return

    lang = db.get_user_language(user_id)
    texts = {
        'uz': '📊 Statistika\n\nDavrni tanlang:',
        'ru': '📊 Статистика\n\nВыберите период:',
        'en': '📊 Statistics\n\nChoose period:'
    }

    await message.answer(
        texts.get(lang, texts['uz']),
        reply_markup=get_statistics_keyboard(user_id)
    )


async def show_weekly_statistics(message: types.Message):
    user_id = message.from_user.id

    if not is_admin(user_id):
        return

    stats = db.get_statistics('week')

    text = f'''📊 Haftalik statistika (oxirgi 7 kun)

👥 Yangi foydalanuvchilar: {stats.get('new_users', 0)}
📬 Yangi murojaatlar: {stats.get('new_applications', 0)}
✅ Javob berilgan: {stats.get('answered', 0)}
⏳ Kutilmoqda: {stats.get('pending', 0)}

📈 Umumiy:
👥 Jami foydalanuvchilar: {stats.get('total_users', 0)}
📬 Jami murojaatlar: {stats.get('total_applications', 0)}

📅 {get_tashkent_now().strftime("%Y-%m-%d %H:%M")}'''

    await message.answer(text, reply_markup=get_admin_keyboard(user_id))


async def show_monthly_statistics(message: types.Message):
    user_id = message.from_user.id

    if not is_admin(user_id):
        return

    stats = db.get_statistics('month')

    text = f'''📊 Oylik statistika (oxirgi 30 kun)

👥 Yangi foydalanuvchilar: {stats.get('new_users', 0)}
📬 Yangi murojaatlar: {stats.get('new_applications', 0)}
✅ Javob berilgan: {stats.get('answered', 0)}
⏳ Kutilmoqda: {stats.get('pending', 0)}

📈 Umumiy:
👥 Jami foydalanuvchilar: {stats.get('total_users', 0)}
📬 Jami murojaatlar: {stats.get('total_applications', 0)}

📅 {get_tashkent_now().strftime("%Y-%m-%d %H:%M")}'''

    await message.answer(text, reply_markup=get_admin_keyboard(user_id))


# ==================== BROADCAST ====================
# ESKI BUTTON-BASED BROADCAST O'CHIRILDI
# Endi /broadcast komandasi ishlatiladi (universal_broadcast.py)
# Eski handler'lar git history'dan topish mumkin


async def add_event_handler(message: types.Message, state: FSMContext):
    user_id = message.from_user.id

    if not is_admin(user_id):
        return

    await message.answer(
        t(user_id, 'enter_event_title'),
        reply_markup=get_cancel_keyboard(user_id)
    )
    await EventCreateState.waiting_for_title.set()


async def process_event_title(message: types.Message, state: FSMContext):
    user_id = message.from_user.id

    if message.text in ['❌ Bekor qilish', '❌ Отмена', '❌ Cancel']:
        await state.finish()
        await message.answer(
            t(user_id, 'admin_panel'),
            reply_markup=get_admin_keyboard(user_id)
        )
        return

    await state.update_data(title=message.text)
    await message.answer(t(user_id, 'enter_event_description'))
    await EventCreateState.waiting_for_description.set()


async def process_event_description(message: types.Message, state: FSMContext):
    user_id = message.from_user.id

    await state.update_data(description=message.text)
    await message.answer(t(user_id, 'enter_event_date'))
    await EventCreateState.waiting_for_date.set()


async def process_event_date(message: types.Message, state: FSMContext):
    user_id = message.from_user.id

    await state.update_data(date=message.text)
    await message.answer(t(user_id, 'enter_event_time'))
    await EventCreateState.waiting_for_time.set()


async def process_event_time(message: types.Message, state: FSMContext):
    user_id = message.from_user.id

    await state.update_data(time=message.text)
    await message.answer(t(user_id, 'enter_event_location'))
    await EventCreateState.waiting_for_location.set()


async def process_event_location(message: types.Message, state: FSMContext):
    user_id = message.from_user.id

    await state.update_data(location=message.text)
    await message.answer(
        t(user_id, 'enter_event_registration_link'),
        reply_markup=get_skip_keyboard(user_id)
    )
    await EventCreateState.waiting_for_registration_link.set()


async def process_event_registration_link(message: types.Message, state: FSMContext):
    user_id = message.from_user.id

    # Check if user wants to skip
    if message.text in ['⏭ O\'tkazib yuborish', '⏭ Пропустить', '⏭ Skip']:
        await state.update_data(registration_link=None)
    else:
        await state.update_data(registration_link=message.text)

    await message.answer(
        t(user_id, 'send_event_image'),
        reply_markup=get_skip_keyboard(user_id)
    )
    await EventCreateState.waiting_for_image.set()


async def process_event_image(message: types.Message, state: FSMContext):
    user_id = message.from_user.id

    image_id = None

    if message.photo:
        image_id = message.photo[-1].file_id
    elif message.document:
        image_id = message.document.file_id
    elif message.text and (message.text.lower() in ['skip', 'yo\'q', 'нет', 'no'] or
                          message.text in ['⏭ O\'tkazib yuborish', '⏭ Пропустить', '⏭ Skip']):
        image_id = None

    data = await state.get_data()

    # Create event with new fields
    db.create_event(
        title=data['title'],
        description=data['description'],
        date=data['date'],
        location=data['location'],
        image_id=image_id,
        time=data.get('time'),
        registration_link=data.get('registration_link')
    )

    await message.answer(
        t(user_id, 'event_created'),
        reply_markup=get_admin_keyboard(user_id)
    )
    await state.finish()


async def manage_events_handler(message: types.Message, state: FSMContext):
    """Admin event management - shows inline keyboard with edit/delete options"""
    user_id = message.from_user.id

    if not is_admin(user_id):
        return

    events = db.get_all_events()

    if not events:
        await message.answer(
            t(user_id, 'no_events'),
            reply_markup=get_admin_keyboard(user_id)
        )
        return

    from keyboards.inline import get_admin_events_keyboard

    await message.answer(
        t(user_id, 'manage_events_title'),
        reply_markup=get_admin_events_keyboard(events)
    )


# Event edit callback handlers
# EDIT FUNKSIYASI O'CHIRILDI - FAQAT DELETE QOLDIRILDI
# async def edit_event_callback(callback: types.CallbackQuery, state: FSMContext):
#     """Handle edit event button click"""
#     user_id = callback.from_user.id
#
#     if not is_admin(user_id):
#         await callback.answer("❌ Ruxsat yo'q")
#         return
#
#     try:
#         event_id = int(callback.data.split('_')[-1])
#         from keyboards.inline import get_event_edit_options_keyboard
#
#         await callback.message.edit_text(
#             t(user_id, 'choose_field_to_edit'),
#             reply_markup=get_event_edit_options_keyboard(event_id)
#         )
#     except Exception as e:
#         await callback.answer(f"❌ Xatolik: {str(e)}")
#
#     await callback.answer()


async def delete_event_callback(callback: types.CallbackQuery):
    """Handle delete event button click - show confirmation"""
    user_id = callback.from_user.id

    if not is_admin(user_id):
        await callback.answer("❌ Ruxsat yo'q")
        return

    try:
        event_id = int(callback.data.split('_')[-1])
        event = db.get_event(event_id)

        if not event:
            await callback.answer(t(user_id, 'event_not_found'))
            return

        from keyboards.inline import get_delete_confirm_keyboard

        await callback.message.edit_text(
            f"{t(user_id, 'confirm_delete_event')}\n\n<b>{event[1]}</b>",
            parse_mode='HTML',
            reply_markup=get_delete_confirm_keyboard(event_id)
        )
    except Exception as e:
        await callback.answer(f"❌ Xatolik: {str(e)}")

    await callback.answer()


async def confirm_delete_event_callback(callback: types.CallbackQuery):
    """Confirm and delete the event"""
    user_id = callback.from_user.id

    if not is_admin(user_id):
        await callback.answer("❌ Ruxsat yo'q")
        return

    try:
        event_id = int(callback.data.split('_')[-1])
        db.delete_event(event_id)

        await callback.message.edit_text(
            t(user_id, 'event_deleted')
        )
        await callback.answer("✅ O'chirildi")
    except Exception as e:
        await callback.answer(f"❌ Xatolik: {str(e)}")


async def admin_back_callback(callback: types.CallbackQuery):
    """Handle admin back button"""
    user_id = callback.from_user.id

    await callback.message.delete()
    await callback.message.answer(
        t(user_id, 'admin_panel'),
        reply_markup=get_admin_keyboard(user_id)
    )
    await callback.answer()


async def admin_manage_events_callback(callback: types.CallbackQuery):
    """Go back to manage events list"""
    user_id = callback.from_user.id

    events = db.get_all_events()

    if not events:
        await callback.message.edit_text(t(user_id, 'no_events'))
        await callback.answer()
        return

    from keyboards.inline import get_admin_events_keyboard

    await callback.message.edit_text(
        t(user_id, 'manage_events_title'),
        reply_markup=get_admin_events_keyboard(events)
    )
    await callback.answer()


# ==================== JADVAL YUKLASH ====================

async def upload_schedule_start(message: types.Message, state: FSMContext):
    """Step 1: Choose faculty for schedule upload"""
    user_id = message.from_user.id

    if not is_admin(user_id):
        return

    await state.finish()

    lang = db.get_user_language(user_id)
    texts = {
        'uz': '📅 Jadval yuklash\n\nFakultetni tanlang:',
        'ru': '📅 Загрузить расписание\n\nВыберите факультет:',
        'en': '📅 Upload Schedule\n\nChoose faculty:'
    }

    await message.answer(
        texts.get(lang, texts['uz']),
        reply_markup=get_faculty_keyboard(user_id)
    )
    await ScheduleUploadState.waiting_for_faculty.set()


async def process_schedule_faculty(message: types.Message, state: FSMContext):
    """Step 2: Choose course"""
    user_id = message.from_user.id

    if message.text in ['⬅️ Orqaga', '⬅️ Назад', '⬅️ Back']:
        await state.finish()
        await message.answer(
            t(user_id, 'admin_panel'),
            reply_markup=get_admin_keyboard(user_id)
        )
        return

    faculty = message.text
    await state.update_data(faculty=faculty)

    lang = db.get_user_language(user_id)
    texts = {
        'uz': f'Fakultet: {faculty}\n\nKursni tanlang:',
        'ru': f'Факультет: {faculty}\n\nВыберите курс:',
        'en': f'Faculty: {faculty}\n\nChoose course:'
    }

    await message.answer(
        texts.get(lang, texts['uz']),
        reply_markup=get_course_keyboard(user_id, faculty)
    )
    await ScheduleUploadState.waiting_for_course.set()


async def process_schedule_course(message: types.Message, state: FSMContext):
    """Step 3: Choose direction or group"""
    user_id = message.from_user.id

    if message.text in ['⬅️ Orqaga', '⬅️ Назад', '⬅️ Back']:
        await message.answer(
            t(user_id, 'choose_faculty'),
            reply_markup=get_faculty_keyboard(user_id)
        )
        await ScheduleUploadState.waiting_for_faculty.set()
        return

    course = message.text
    data = await state.get_data()
    faculty = data.get('faculty')
    await state.update_data(course=course)

    # Check if this is Yurisprudensiya (groups directly) or other faculty (directions first)
    from config import FACULTIES
    lang = db.get_user_language(user_id)
    faculties_lang = FACULTIES.get(lang, FACULTIES['uz'])

    texts = {
        'uz': f'Fakultet: {faculty}\nKurs: {course}\n\n',
        'ru': f'Факультет: {faculty}\nКурс: {course}\n\n',
        'en': f'Faculty: {faculty}\nCourse: {course}\n\n'
    }

    if faculty in faculties_lang and course in faculties_lang[faculty]:
        course_data = faculties_lang[faculty][course]

        if isinstance(course_data, list):
            # Yurisprudensiya - show groups directly
            texts['uz'] += 'Guruhni tanlang:'
            texts['ru'] += 'Выберите группу:'
            texts['en'] += 'Choose group:'

            await message.answer(
                texts.get(lang, texts['uz']),
                reply_markup=get_direction_keyboard(user_id, faculty, course)
            )
            await ScheduleUploadState.waiting_for_direction.set()
        else:
            # Other faculty - show directions
            texts['uz'] += 'Yo\'nalishni tanlang:'
            texts['ru'] += 'Выберите направление:'
            texts['en'] += 'Choose direction:'

            await message.answer(
                texts.get(lang, texts['uz']),
                reply_markup=get_direction_keyboard(user_id, faculty, course)
            )
            await ScheduleUploadState.waiting_for_direction.set()


async def process_schedule_direction(message: types.Message, state: FSMContext):
    """Step 4: Choose group or upload image"""
    user_id = message.from_user.id

    if message.text in ['⬅️ Orqaga', '⬅️ Назад', '⬅️ Back']:
        data = await state.get_data()
        faculty = data.get('faculty')
        await message.answer(
            t(user_id, 'choose_course'),
            reply_markup=get_course_keyboard(user_id, faculty)
        )
        await ScheduleUploadState.waiting_for_course.set()
        return

    direction_or_group = message.text
    data = await state.get_data()
    faculty = data.get('faculty')
    course = data.get('course')

    # Check if this is Yurisprudensiya or other faculty
    from config import FACULTIES
    lang = db.get_user_language(user_id)
    faculties_lang = FACULTIES.get(lang, FACULTIES['uz'])

    if faculty in faculties_lang and course in faculties_lang[faculty]:
        course_data = faculties_lang[faculty][course]

        if isinstance(course_data, list):
            # Yurisprudensiya - direction_or_group is actually a group, ask for image
            group = direction_or_group
            await state.update_data(group=group, direction='')

            texts = {
                'uz': f'Fakultet: {faculty}\nKurs: {course}\nGuruh: {group}\n\nJadval rasmini yuboring:',
                'ru': f'Факультет: {faculty}\nКурс: {course}\nГруппа: {group}\n\nОтправьте изображение расписания:',
                'en': f'Faculty: {faculty}\nCourse: {course}\nGroup: {group}\n\nSend schedule image:'
            }

            await message.answer(
                texts.get(lang, texts['uz']),
                reply_markup=get_cancel_keyboard(user_id)
            )
            await ScheduleUploadState.waiting_for_image.set()
        else:
            # Other faculty - direction_or_group is a direction, show groups
            direction = direction_or_group
            await state.update_data(direction=direction)

            groups = course_data.get(direction, [])

            texts = {
                'uz': f'Fakultet: {faculty}\nKurs: {course}\nYo\'nalish: {direction}\n\nGuruhni tanlang:',
                'ru': f'Факультет: {faculty}\nКурс: {course}\nНаправление: {direction}\n\nВыберите группу:',
                'en': f'Faculty: {faculty}\nCourse: {course}\nDirection: {direction}\n\nChoose group:'
            }

            await message.answer(
                texts.get(lang, texts['uz']),
                reply_markup=get_group_keyboard(user_id, groups)
            )
            await ScheduleUploadState.waiting_for_group.set()


async def process_schedule_group(message: types.Message, state: FSMContext):
    """Step 5: Request image for group"""
    user_id = message.from_user.id

    if message.text in ['⬅️ Orqaga', '⬅️ Назад', '⬅️ Back']:
        data = await state.get_data()
        faculty = data.get('faculty')
        course = data.get('course')
        await message.answer(
            t(user_id, 'choose_direction'),
            reply_markup=get_direction_keyboard(user_id, faculty, course)
        )
        await ScheduleUploadState.waiting_for_direction.set()
        return

    group = message.text
    data = await state.get_data()
    faculty = data.get('faculty')
    course = data.get('course')
    direction = data.get('direction')

    await state.update_data(group=group)

    lang = db.get_user_language(user_id)
    texts = {
        'uz': f'Fakultet: {faculty}\nKurs: {course}\nYo\'nalish: {direction}\nGuruh: {group}\n\nJadval rasmini yuboring:',
        'ru': f'Факультет: {faculty}\nКурс: {course}\nНаправление: {direction}\nГруппа: {group}\n\nОтправьте изображение расписания:',
        'en': f'Faculty: {faculty}\nCourse: {course}\nDirection: {direction}\nGroup: {group}\n\nSend schedule image:'
    }

    await message.answer(
        texts.get(lang, texts['uz']),
        reply_markup=get_cancel_keyboard(user_id)
    )
    await ScheduleUploadState.waiting_for_image.set()


async def process_schedule_image(message: types.Message, state: FSMContext):
    """Step 6: Save schedule image"""
    user_id = message.from_user.id

    if message.text and message.text in ['❌ Bekor qilish', '❌ Отмена', '❌ Cancel']:
        await state.finish()
        await message.answer(
            t(user_id, 'admin_panel'),
            reply_markup=get_admin_keyboard(user_id)
        )
        return

    # Get image file_id
    image_id = None
    if message.photo:
        image_id = message.photo[-1].file_id
    elif message.document:
        image_id = message.document.file_id
    else:
        await message.answer('❌ Iltimos, rasm yuboring!')
        return

    data = await state.get_data()
    faculty = data.get('faculty')
    direction = data.get('direction', '')
    course = data.get('course')
    group = data.get('group')

    # Save to database
    try:
        # Use database method with direction parameter
        # We need to insert or update schedule
        conn = db.get_connection()
        c = conn.cursor()

        # Check if schedule exists
        c.execute(
            "SELECT id FROM schedules WHERE faculty=? AND direction=? AND course=? AND group_name=?",
            (faculty, direction, course, group)
        )
        existing = c.fetchone()

        if existing:
            # Update
            c.execute(
                "UPDATE schedules SET image_id=? WHERE id=?",
                (image_id, existing[0])
            )
        else:
            # Insert
            c.execute(
                '''INSERT INTO schedules (faculty, direction, course, group_name, image_id, created_at)
                   VALUES (?,?,?,?,?,?)''',
                (faculty, direction, course, group, image_id, get_tashkent_now().strftime("%Y-%m-%d %H:%M:%S"))
            )

        conn.commit()
        conn.close()

        lang = db.get_user_language(user_id)

        # Prepare direction text
        dir_uz = direction if direction else "Yo'q"
        dir_ru = direction if direction else "Нет"
        dir_en = direction if direction else "None"

        texts = {
            'uz': f'✅ Jadval muvaffaqiyatli saqlandi!\n\nFakultet: {faculty}\nKurs: {course}\nYo\'nalish: {dir_uz}\nGuruh: {group}',
            'ru': f'✅ Расписание успешно сохранено!\n\nФакультет: {faculty}\nКурс: {course}\nНаправление: {dir_ru}\nГруппа: {group}',
            'en': f'✅ Schedule saved successfully!\n\nFaculty: {faculty}\nCourse: {course}\nDirection: {dir_en}\nGroup: {group}'
        }

        await message.answer_photo(
            image_id,
            caption=texts.get(lang, texts['uz']),
            reply_markup=get_admin_keyboard(user_id)
        )
    except Exception as e:
        logger.error(f'Error saving schedule: {e}')
        await message.answer(
            f'❌ Xatolik yuz berdi: {e}',
            reply_markup=get_admin_keyboard(user_id)
        )

    await state.finish()


# ==================== KONTRAKT YUKLASH ====================

async def upload_contract_start(message: types.Message, state: FSMContext):
    """Start contract upload - request Excel file"""
    user_id = message.from_user.id

    if not is_admin(user_id):
        return

    await state.finish()

    lang = db.get_user_language(user_id)
    texts = {
        'uz': '💼 <b>Kontrakt ma\'lumotlarini yuklash</b>\n\nExcel faylni yuboring.',
        'ru': '💼 <b>Загрузка данных контрактов</b>\n\nОтправьте Excel файл.',
        'en': '💼 <b>Upload Contract Data</b>\n\nSend an Excel file.'
    }

    # Show current statistics
    contracts_count = db.get_contracts_count()
    last_upload = db.get_last_contract_upload_date()

    stats_text = ''
    if contracts_count > 0:
        stats_uz = f'\n\n📊 Hozirgi holat:\n• Kontrakt ma\'lumotlari: {contracts_count} ta\n• Oxirgi yangilanish: {last_upload}'
        stats_ru = f'\n\n📊 Текущее состояние:\n• Данные контрактов: {contracts_count} шт.\n• Последнее обновление: {last_upload}'
        stats_en = f'\n\n📊 Current status:\n• Contract data: {contracts_count} records\n• Last update: {last_upload}'

        stats_text = stats_uz if lang == 'uz' else stats_ru if lang == 'ru' else stats_en

    await message.answer(
        texts.get(lang, texts['uz']) + stats_text,
        reply_markup=get_cancel_keyboard(user_id),
        parse_mode='HTML'
    )

    from states.forms import ContractUploadState
    await ContractUploadState.waiting_for_excel.set()


async def process_contract_excel(message: types.Message, state: FSMContext):
    """Process uploaded Excel file with contract data"""
    user_id = message.from_user.id
    lang = db.get_user_language(user_id)

    if message.text and message.text in ['❌ Bekor qilish', '❌ Отмена', '❌ Cancel']:
        await state.finish()
        await message.answer(
            t(user_id, 'admin_panel'),
            reply_markup=get_admin_keyboard(user_id)
        )
        return

    # Check if document is provided
    if not message.document:
        await message.answer('❌ Iltimos, Excel faylni yuboring!')
        return

    # Check file extension
    file_name = message.document.file_name
    if not (file_name.endswith('.xlsx') or file_name.endswith('.xls')):
        texts = {
            'uz': '❌ Noto\'g\'ri fayl formati! Iltimos, .xlsx yoki .xls formatdagi faylni yuboring.',
            'ru': '❌ Неверный формат файла! Пожалуйста, отправьте файл в формате .xlsx или .xls.',
            'en': '❌ Invalid file format! Please send a file in .xlsx or .xls format.'
        }
        await message.answer(texts.get(lang, texts['uz']))
        return

    # Download file
    try:
        file = await message.bot.get_file(message.document.file_id)
        file_path = f"/tmp/contracts_{user_id}_{datetime.now().timestamp()}.xlsx"
        await message.bot.download_file(file.file_path, file_path)

        # Parse Excel file
        try:
            import openpyxl
        except ImportError:
            await message.answer('❌ Server xatolik: openpyxl kutubxonasi o\'rnatilmagan')
            await state.finish()
            return

        await message.answer('⏳ Fayl qayta ishlanmoqda...')

        workbook = openpyxl.load_workbook(file_path, data_only=True)
        sheet = workbook.active

        logger.info(f"Excel file loaded. Sheet: {sheet.title}, Max row: {sheet.max_row}, Max col: {sheet.max_column}")

        contracts_data = []
        errors = []

        # Helper function to parse amounts
        def parse_amount(value):
            """Parse amount from various formats"""
            if value is None:
                return 0.0
            try:
                # If already a number
                if isinstance(value, (int, float)):
                    return float(value)
                # If string, remove common separators and spaces
                if isinstance(value, str):
                    cleaned = value.replace(',', '').replace(' ', '').replace('\xa0', '').strip()
                    return float(cleaned) if cleaned else 0.0
                return 0.0
            except Exception as e:
                logger.warning(f"Could not parse amount '{value}': {e}")
                return 0.0

        # Skip header row, start from row 2
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not row or not any(row):  # Skip empty rows
                continue

            try:
                # Parse row data
                # row[0] = № (serial number)
                # row[1] = Full Name of student
                # row[2] = Pasport raqami
                # row[3] = JSHSHIR-kod
                # row[4] = Talaba kursi
                # row[5] = Суммаси (total_amount)
                # row[6] = холатига тўлов (paid_amount)
                # row[7] = қолдиқ сумма (remaining_amount)

                full_name = str(row[1]).strip() if row[1] else None
                passport_series = str(row[2]).strip() if row[2] else None
                jshshir = str(row[3]).strip() if row[3] else None
                course = str(row[4]).strip() if row[4] else None

                # Parse amounts using helper function
                total_amount = parse_amount(row[5])
                paid_amount = parse_amount(row[6])
                remaining_amount = parse_amount(row[7])

                if not passport_series or not full_name:
                    errors.append(f"Qator {row_idx}: Pasport yoki ism kiritilmagan")
                    continue

                contracts_data.append({
                    'passport_series': passport_series,
                    'full_name': full_name,
                    'jshshir': jshshir,
                    'course': course,
                    'total_amount': total_amount,
                    'paid_amount': paid_amount,
                    'remaining_amount': remaining_amount
                })

            except Exception as e:
                errors.append(f"Qator {row_idx}: {str(e)}")
                continue

        # Log parsing results
        logger.info(f"Parsed {len(contracts_data)} contracts. Errors: {len(errors)}")

        # Save to database
        if contracts_data:
            inserted_count = db.save_contracts_from_excel(contracts_data, file_name)

            success_texts = {
                'uz': f'''✅ <b>Kontrakt ma'lumotlari yuklandi!</b>

📊 Yuklangan: {inserted_count} ta
📁 Fayl: {file_name}
📅 Sana: {get_tashkent_now().strftime("%Y-%m-%d %H:%M")}''',
                'ru': f'''✅ <b>Данные контрактов загружены!</b>

📊 Загружено: {inserted_count} шт.
📁 Файл: {file_name}
📅 Дата: {get_tashkent_now().strftime("%Y-%m-%d %H:%M")}''',
                'en': f'''✅ <b>Contract data uploaded!</b>

📊 Uploaded: {inserted_count} records
📁 File: {file_name}
📅 Date: {get_tashkent_now().strftime("%Y-%m-%d %H:%M")}'''
            }

            result_text = success_texts.get(lang, success_texts['uz'])

            if errors:
                error_summary = '\n\n⚠️ Xatoliklar:\n' + '\n'.join(errors[:10])
                if len(errors) > 10:
                    error_summary += f'\n... va yana {len(errors) - 10} ta xatolik'
                result_text += error_summary

            await message.answer(
                result_text,
                reply_markup=get_admin_keyboard(user_id),
                parse_mode='HTML'
            )
        else:
            await message.answer(
                '❌ Hech qanday ma\'lumot yuklanmadi. Faylni tekshiring.',
                reply_markup=get_admin_keyboard(user_id)
            )

        # Clean up temp file
        import os
        if os.path.exists(file_path):
            os.remove(file_path)

    except Exception as e:
        logger.error(f'Error processing contract Excel: {e}')
        await message.answer(
            f'❌ Xatolik yuz berdi: {str(e)}',
            reply_markup=get_admin_keyboard(user_id)
        )

    await state.finish()


# EDIT FUNKSIYALARI O'CHIRILDI - FAQAT DELETE QOLDIRILDI
# async def edit_field_callback(callback: types.CallbackQuery, state: FSMContext):
#     """Handle field edit button click"""
#     user_id = callback.from_user.id
#
#     if not is_admin(user_id):
#         await callback.answer("❌ Ruxsat yo'q")
#         return
#
#     # Parse callback data: edit_<field>_<event_id>
#     parts = callback.data.split('_')
#     field = parts[1]
#     event_id = int(parts[2])
#
#     # Store event_id and field in state
#     await state.update_data(event_id=event_id, field=field)
#
#     # Get field name in user's language
#     field_names = {
#         'title': 'nomini',
#         'desc': 'tavsifini',
#         'date': 'sanasini',
#         'time': 'vaqtini',
#         'location': 'manzilini',
#         'link': 'havolasini',
#         'image': 'rasmini'
#     }
#
#     await callback.message.edit_text(
#         f"Tadbirning {field_names.get(field, field)} yangi qiymatini kiriting:",
#         reply_markup=get_cancel_keyboard(user_id)
#     )
#
#     from states.forms import EventEditState
#     await EventEditState.waiting_for_field_value.set()
#     await callback.answer()
#
#
# async def process_field_edit(message: types.Message, state: FSMContext):
#     """Process the new field value"""
#     user_id = message.from_user.id
#
#     if message.text in ['❌ Bekor qilish', '❌ Отмена', '❌ Cancel']:
#         await state.finish()
#         await message.answer(
#             t(user_id, 'admin_panel'),
#             reply_markup=get_admin_keyboard(user_id)
#         )
#         return
#
#     data = await state.get_data()
#     event_id = data.get('event_id')
#     field = data.get('field')
#
#     # Get current event data
#     event = db.get_event(event_id)
#     if not event:
#         await message.answer(t(user_id, 'event_not_found'))
#         await state.finish()
#         return
#
#     # Update the specific field
#     if field == 'title':
#         db.update_event(event_id, message.text, event[2], event[3], event[4], event[5], event[6], event[7])
#     elif field == 'desc':
#         db.update_event(event_id, event[1], message.text, event[3], event[4], event[5], event[6], event[7])
#     elif field == 'date':
#         db.update_event(event_id, event[1], event[2], message.text, event[4], event[5], event[6], event[7])
#     elif field == 'time':
#         db.update_event(event_id, event[1], event[2], event[3], message.text, event[5], event[6], event[7])
#     elif field == 'location':
#         db.update_event(event_id, event[1], event[2], event[3], event[4], message.text, event[6], event[7])
#     elif field == 'link':
#         db.update_event(event_id, event[1], event[2], event[3], event[4], event[5], message.text, event[7])
#     elif field == 'image':
#         if message.photo:
#             image_id = message.photo[-1].file_id
#             db.update_event(event_id, event[1], event[2], event[3], event[4], event[5], event[6], image_id)
#         elif message.document:
#             image_id = message.document.file_id
#             db.update_event(event_id, event[1], event[2], event[3], event[4], event[5], event[6], image_id)
#
#     await message.answer(
#         "✅ Tadbir yangilandi!",
#         reply_markup=get_admin_keyboard(user_id)
#     )
#     await state.finish()


def register_admin_handlers(dp: Dispatcher):
    # Admin panel asosiy
    dp.register_message_handler(
        admin_panel_handler,
        lambda message: message.text in ['👨‍💼 Admin panel', '👨‍💼 Админ панель', '👨‍💼 Admin Panel'],
        state='*'
    )

    # Yangi murojaatlar
    dp.register_message_handler(
        view_new_applications_handler,
        lambda message: message.text in ['📬 Yangi murojaatlar', '📬 Новые обращения', '📬 New Applications'] and is_admin(
            message.from_user.id)
    )

    # Javob berilgan murojaatlar
    dp.register_message_handler(
        view_answered_applications_handler,
        lambda message: message.text in ['✅ Javob berilganlar', '✅ Отвеченные', '✅ Answered'] and is_admin(
            message.from_user.id)
    )

    # Murojaatga javob
    dp.register_message_handler(
        reply_command_handler,
        lambda message: message.text.startswith('/reply_'),
        state='*'
    )
    dp.register_message_handler(process_admin_reply, state=AdminReplyState.waiting_for_reply)

    # Statistika
    dp.register_message_handler(
        statistics_handler,
        lambda message: message.text in ['📊 Statistika', '📊 Статистика', '📊 Statistics'] and is_admin(
            message.from_user.id)
    )
    dp.register_message_handler(
        show_weekly_statistics,
        lambda message: message.text in ['📅 Haftalik', '📅 Недельная', '📅 Weekly'] and is_admin(message.from_user.id)
    )
    dp.register_message_handler(
        show_monthly_statistics,
        lambda message: message.text in ['📆 Oylik', '📆 Месячная', '📆 Monthly'] and is_admin(message.from_user.id)
    )

    # Broadcast - ESKI BUTTON-BASED METHOD O'CHIRILDI
    # Endi /broadcast komandasi ishlatiladi (universal_broadcast.py)
    # dp.register_message_handler(
    #     broadcast_start_handler,
    #     lambda message: message.text in ['📢 Broadcast', '📢 Рассылка'] and is_admin(message.from_user.id),
    #     state='*'
    # )
    # dp.register_message_handler(
    #     broadcast_process_message,
    #     content_types=['text', 'photo', 'video', 'document'],
    #     state=BroadcastState.waiting_for_content
    # )
    # dp.register_message_handler(
    #     broadcast_confirm,
    #     state=BroadcastState.waiting_for_confirmation
    # )

    dp.register_message_handler(process_admin_reply, state=AdminReplyState.waiting_for_reply)

    # ESKI USUL - O'CHIRILDI (event_quick_create.py ishlatiladi)
    # dp.register_message_handler(
    #     add_event_handler,
    #     lambda message: message.text in ['➕ Tadbir qo\'shish', '➕ Добавить мероприятие'] and is_admin(
    #         message.from_user.id),
    #     state='*'
    # )
    # dp.register_message_handler(process_event_title, state=EventCreateState.waiting_for_title)
    # dp.register_message_handler(process_event_description, state=EventCreateState.waiting_for_description)
    # dp.register_message_handler(process_event_date, state=EventCreateState.waiting_for_date)
    # dp.register_message_handler(process_event_time, state=EventCreateState.waiting_for_time)
    # dp.register_message_handler(process_event_location, state=EventCreateState.waiting_for_location)
    # dp.register_message_handler(process_event_registration_link, state=EventCreateState.waiting_for_registration_link)
    # dp.register_message_handler(
    #     process_event_image,
    #     content_types=['text', 'photo', 'document'],
    #     state=EventCreateState.waiting_for_image
    # )

    dp.register_message_handler(
        manage_events_handler,
        lambda message: message.text in ['📝 Tadbirlarni boshqarish', '📝 Управление мероприятиями'] and is_admin(
            message.from_user.id),
        state='*'
    )

    # Event delete callback handlers (Edit handlers o'chirildi)
    # dp.register_callback_query_handler(
    #     edit_event_callback,
    #     lambda c: c.data.startswith('edit_event_') and is_admin(c.from_user.id),
    #     state='*'
    # )
    dp.register_callback_query_handler(
        delete_event_callback,
        lambda c: c.data.startswith('delete_event_') and is_admin(c.from_user.id)
    )
    dp.register_callback_query_handler(
        confirm_delete_event_callback,
        lambda c: c.data.startswith('confirm_delete_') and is_admin(c.from_user.id)
    )
    dp.register_callback_query_handler(
        admin_back_callback,
        lambda c: c.data == 'admin_back' and is_admin(c.from_user.id)
    )
    dp.register_callback_query_handler(
        admin_manage_events_callback,
        lambda c: c.data == 'admin_manage_events' and is_admin(c.from_user.id)
    )
    # dp.register_callback_query_handler(
    #     edit_field_callback,
    #     lambda c: c.data.startswith('edit_') and not c.data.startswith('edit_event_') and is_admin(c.from_user.id),
    #     state='*'
    # )

    # Event edit state handler - O'CHIRILDI
    # from states.forms import EventEditState
    # dp.register_message_handler(
    #     process_field_edit,
    #     content_types=['text', 'photo', 'document'],
    #     state=EventEditState.waiting_for_field_value
    # )

    # Schedule upload handlers
    dp.register_message_handler(
        upload_schedule_start,
        lambda message: message.text in ['📅 Jadval yuklash', '📅 Загрузить расписание', '📅 Upload Schedule'] and is_admin(
            message.from_user.id),
        state='*'
    )
    dp.register_message_handler(process_schedule_faculty, state=ScheduleUploadState.waiting_for_faculty)
    dp.register_message_handler(process_schedule_course, state=ScheduleUploadState.waiting_for_course)
    dp.register_message_handler(process_schedule_direction, state=ScheduleUploadState.waiting_for_direction)
    dp.register_message_handler(process_schedule_group, state=ScheduleUploadState.waiting_for_group)
    dp.register_message_handler(
        process_schedule_image,
        content_types=['photo', 'document'],
        state=ScheduleUploadState.waiting_for_image
    )

    # Contract upload handlers
    dp.register_message_handler(
        upload_contract_start,
        lambda message: message.text in ['💼 Kontrakt yuklash', '💼 Загрузить контракты', '💼 Upload Contracts'] and is_admin(
            message.from_user.id),
        state='*'
    )
    from states.forms import ContractUploadState
    dp.register_message_handler(
        process_contract_excel,
        content_types=['document', 'text'],
        state=ContractUploadState.waiting_for_excel
    )